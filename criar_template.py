#!/usr/bin/env python3
"""
Cria um template em branco da Solicitação de Emprego.
Preserva toda a estrutura/formatação e apaga apenas os dados do candidato.

Uso:
    python criar_template.py --origem "SOLICITAÇÃO DE EMPREGO ULYSSES CORREA CESAR.xlsx"
"""

import sys
import shutil
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Erro: pip install openpyxl")
    sys.exit(1)

# Células que contêm dados do candidato (não labels)
DATA_CELLS = [
    "A3",   # Nome
    "B4",   # Endereço
    "A5",   # CEP (substituído pelo label fixo)
    "C5",   # Bairro
    "F5",   # Estado residência
    "B6",   # Telefone
    "A8",   # Data nascimento
    "B8",   # Idade
    "C8",   # Local nascimento
    "F8",   # Estado natal
    "A10",  # Estado civil
    "B10",  # Nacionalidade
    "D10",  # Escolaridade
    "B11",  # Função
    "A13",  # Experiência sim/não
    "B13",  # Tempo experiência
    "D13",  # FIBRA sim
    "F13",  # FIBRA não
    "A14",  # Nome pai
    "A15",  # Nome mãe
    # Beneficiários
    "A17", "E17", "A18",
    "A19", "E19",
    "A20", "E20",
    "A21", "E21",
    # Emprego anterior
    "B23",  # Empresa
    "B24",  # Endereço empresa
    "B25",  # Cargo
    "B26",  # Data admissão
    "D26",  # Data demissão
    "B27",  # Motivo saída
    "B28",  # Observação
    # Carteira profissional
    "H7",   # Número carteira
    "J7",   # Série carteira
    "H9",   # Estado emissor carteira
    "J9",   # Data emissão carteira
    # RG
    "H12",  # Número RG
    "H14",  # Estado emissor RG
    "J14",  # Data emissão RG
    # CPF
    "H17",  # Número CPF
    # Título eleitor
    "H20",  # Número título
    "H22",  # Zona
    "J22",  # Seção
    # Reservista
    "H25",  # Número reservista
    "H27",  # Série reservista
    "J27",  # Categoria reservista
    # PIS
    "H30",  # Número PIS
    # Dados bancários
    "B32",  # Banco
    "E32",  # Agência
    "B33",  # Conta corrente
    # Mês e obra (deixa em branco; mês é preenchido automaticamente)
    "H4",   # Mês
]

# Células de label fixo que devem ser restauradas após limpar
FIXED_LABELS = {
    "A3":  "NOME: ",
    "A5":  "CEP:         ",
    "A13": "Sim:        Não:       ",
    "B13": "Tempo:  ",
    "D13": "Sim:",
    "F13": "Não:",
    "A14": "PAI: ",
    "A15": "MÃE:  ",
    "A17": "Nome: ",
    "A18": "CPF:  ",
    "A19": "Nome:",
    "A20": "Nome:",
    "A21": "Nome:",
    "E17": "Parentesco:",
    "E19": "Parentesco:",
    "E20": "Parentesco:",
    "E21": "Parentesco:",
    "B28": "OBSERVAÇÃO:",
}


def get_merge_top_left(ws, cell_coord):
    """Retorna a célula top-left do range mesclado que contém cell_coord, ou None se não for mesclada."""
    for merge_range in ws.merged_cells.ranges:
        if cell_coord in merge_range:
            # Retorna coordenada top-left
            top_row = merge_range.min_row
            top_col = merge_range.min_col
            return ws.cell(row=top_row, column=top_col)
    return None


def safe_set_cell(ws, cell_coord, value):
    """Define valor de célula de forma segura, mesmo que seja parte de merge."""
    # Verifica se a célula é parte de um merge
    top_left = get_merge_top_left(ws, cell_coord)
    if top_left:
        # Só permite escrever se for a célula top-left
        from openpyxl.utils.cell import coordinate_from_string
        from openpyxl.utils import column_index_from_string
        col_str, row_idx = coordinate_from_string(cell_coord)
        col_idx = column_index_from_string(col_str)
        if top_left.row == row_idx and top_left.column == col_idx:
            top_left.value = value
        # Se não for top-left, ignora silenciosamente
    else:
        ws[cell_coord] = value


def create_blank_template(origem: str, destino: str):
    shutil.copy2(origem, destino)
    wb = openpyxl.load_workbook(destino)
    ws = wb.active

    for cell_coord in DATA_CELLS:
        safe_set_cell(ws, cell_coord, None)

    for cell_coord, label in FIXED_LABELS.items():
        safe_set_cell(ws, cell_coord, label)

    wb.save(destino)
    print(f"Template em branco criado: {destino}")


def main():
    parser = argparse.ArgumentParser(description="Cria template em branco da Solicitação de Emprego")
    parser.add_argument("--origem", "-o", required=True, help="Arquivo Excel de origem (com dados de exemplo)")
    parser.add_argument("--destino", "-d", default="template_solicitacao.xlsx", help="Arquivo de destino")
    args = parser.parse_args()

    if not Path(args.origem).exists():
        print(f"Erro: arquivo não encontrado: {args.origem}")
        sys.exit(1)

    create_blank_template(args.origem, args.destino)


if __name__ == "__main__":
    main()
