#!/usr/bin/env python3
"""
Sistema de Preenchimento de Solicitação de Emprego - FIBRA
Usa Google Gemini (gratuito) para ler documentos e preencher o Excel.

Uso:
    python solicitacao_emprego.py --documentos ./docs_joao/ --api-key SUA_CHAVE

Múltiplas chaves (para não esgotar a cota):
    python solicitacao_emprego.py --documentos ./docs/ --api-key CHAVE1 --api-key CHAVE2

Nomeie os arquivos de documentos assim (aceita .jpg, .jpeg, .png, .pdf, .webp):
    rg_frente.jpg               RG frente
    rg_verso.jpg                RG verso
    cpf_frente.jpg              CPF frente
    cpf_verso.jpg               CPF verso (opcional)
    carteira_trabalho.jpg       Carteira de Trabalho (CTPS)
    carteira_trabalho_digital   Carteira de Trabalho Digital
    cnh.jpg                     CNH
    comprovante_residencia      Comprovante de Residência
    curriculo                   Currículo (PDF ou imagem)
    foto_3x4.jpg                Foto 3x4
    historico_escolar           Histórico Escolar
    certificado_conclusao       Certificado de Conclusão
    titulo_eleitor.jpg          Título de Eleitor
    certificado_reservista.jpg  Certificado de Reservista
    cartao_vacina_frente.jpg    Cartão de Vacina frente
    cartao_vacina_verso.jpg     Cartão de Vacina verso
"""

import os
import sys
import json
import argparse
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from google import genai
    from google.genai import types as gtypes
except ImportError:
    print("Erro: biblioteca 'google-genai' não instalada.")
    print("Execute: pip install google-genai")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Erro: biblioteca 'openpyxl' não instalada.")
    print("Execute: pip install openpyxl")
    sys.exit(1)

# ============================================================
# GERENCIADOR DE CHAVES GEMINI
# ============================================================

DAILY_LIMIT = 1500
STOP_AT_REMAINING = 3   # Para de usar a chave quando restar esse número

KEYS_FILE = Path(__file__).parent / "gemini_keys.json"


class GeminiKeyManager:
    """Gerencia múltiplas chaves Gemini com controle de cota diária."""

    def __init__(self, env_keys: list = None):
        self.data = self._load()
        self._reset_if_new_day()
        # Auto-carrega chaves de variáveis de ambiente
        for key in (env_keys or []):
            if key:
                self.add_key(key)
        env_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
        if env_key:
            self.add_key(env_key)

    def _load(self) -> dict:
        if KEYS_FILE.exists():
            with open(KEYS_FILE, encoding="utf-8") as f:
                return json.load(f)
        return {"keys": [], "current_index": 0}

    def _save(self):
        with open(KEYS_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2)

    def _today(self) -> str:
        return datetime.today().strftime("%Y-%m-%d")

    def _reset_if_new_day(self):
        today = self._today()
        changed = False
        for entry in self.data["keys"]:
            if entry.get("date") != today:
                entry["date"] = today
                entry["requests_today"] = 0
                changed = True
        if changed:
            self._save()

    def add_key(self, api_key: str) -> bool:
        api_key = api_key.strip()
        if not api_key:
            return False
        for entry in self.data["keys"]:
            if entry["key"] == api_key:
                return False  # já existe
        self.data["keys"].append({
            "key": api_key,
            "date": self._today(),
            "requests_today": 0,
        })
        self._save()
        return True

    def has_keys(self) -> bool:
        return len(self.data["keys"]) > 0

    def _current_entry(self) -> Optional[dict]:
        idx = self.data.get("current_index", 0)
        if 0 <= idx < len(self.data["keys"]):
            return self.data["keys"][idx]
        return None

    def _remaining(self, entry: dict) -> int:
        return DAILY_LIMIT - entry.get("requests_today", 0)

    def _try_advance(self) -> bool:
        """Tenta avançar para a próxima chave disponível. Retorna True se encontrou."""
        total = len(self.data["keys"])
        start = self.data.get("current_index", 0)
        for offset in range(1, total + 1):
            idx = (start + offset) % total
            entry = self.data["keys"][idx]
            if self._remaining(entry) > STOP_AT_REMAINING:
                self.data["current_index"] = idx
                self._save()
                return True
        return False

    def get_active_key(self) -> Optional[str]:
        """
        Retorna a chave ativa. Se estiver no limite, tenta avançar.
        Retorna None se todas estiverem esgotadas.
        """
        if not self.data["keys"]:
            return None

        entry = self._current_entry()
        if entry is None:
            self.data["current_index"] = 0
            entry = self._current_entry()

        remaining = self._remaining(entry)

        if remaining <= STOP_AT_REMAINING:
            # Tenta próxima chave
            if self._try_advance():
                entry = self._current_entry()
                remaining = self._remaining(entry)
            else:
                return None  # todas esgotadas

        return entry["key"]

    def consume(self):
        """Registra 1 requisição na chave atual."""
        entry = self._current_entry()
        if entry:
            entry["requests_today"] = entry.get("requests_today", 0) + 1
            entry["date"] = self._today()
            self._save()

    def warn_if_low(self):
        """Exibe aviso se restarem poucas requisições na chave atual."""
        entry = self._current_entry()
        if not entry:
            return
        remaining = self._remaining(entry)
        if remaining <= STOP_AT_REMAINING:
            short = entry["key"][-8:]
            print(f"\n{'!'*60}")
            print(f"  ATENÇÃO: chave ...{short} chegou ao limite.")
            print(f"  Restam {remaining} requisição(ões) reservadas como buffer.")
            print(f"{'!'*60}\n")
        elif remaining <= 20:
            short = entry["key"][-8:]
            print(f"\n[!] Aviso: apenas {remaining} requisições restantes na chave ...{short}")

    def request_new_key_interactive(self) -> bool:
        """Pausa e pede ao usuário uma nova chave Gemini."""
        print("\n" + "=" * 60)
        print("  TODAS AS CHAVES GEMINI ESTÃO NO LIMITE DE HOJE")
        print("=" * 60)
        self.print_status()
        print()
        print("Para obter uma nova chave gratuita:")
        print("  1. Acesse: https://aistudio.google.com/app/apikey")
        print("  2. Faça login com outra conta Google")
        print("  3. Clique em 'Create API key'")
        print("  4. Cole a chave abaixo")
        print()
        print("(Pressione Enter sem digitar para cancelar o processamento)")
        new_key = input("Nova chave Gemini: ").strip()
        if new_key:
            added = self.add_key(new_key)
            if added:
                # Aponta para a nova chave
                self.data["current_index"] = len(self.data["keys"]) - 1
                self._save()
                print(f"[OK] Chave ...{new_key[-8:]} adicionada.\n")
                return True
            else:
                print("[!] Chave já cadastrada.")
        return False

    def print_status(self):
        if not self.data["keys"]:
            print("  Nenhuma chave cadastrada.")
            return
        current_idx = self.data.get("current_index", 0)
        for i, entry in enumerate(self.data["keys"]):
            remaining = self._remaining(entry)
            mark = " <- ativa" if i == current_idx else ""
            short = entry["key"][-8:]
            bar_used = min(20, 20 - int(remaining / DAILY_LIMIT * 20))
            bar = "#" * bar_used + "-" * (20 - bar_used)
            print(f"  Chave ...{short}{mark}")
            print(f"    [{bar}] {entry.get('requests_today', 0)}/{DAILY_LIMIT} usadas, {remaining} restantes")


# ============================================================
# CONSTANTES E CONFIGURAÇÃO
# ============================================================

MESES_PT = {
    1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
    5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
    9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
}

DOCUMENT_TYPES = {
    "rg_frente":                 "RG (Frente)",
    "rg_verso":                  "RG (Verso)",
    "cpf_frente":                "CPF (Frente)",
    "cpf_verso":                 "CPF (Verso)",
    "carteira_trabalho":         "Carteira de Trabalho (CTPS)",
    "carteira_trabalho_digital": "Carteira de Trabalho Digital",
    "cnh":                       "CNH (Carteira de Habilitação)",
    "comprovante_residencia":    "Comprovante de Residência",
    "curriculo":                 "Currículo",
    "foto_3x4":                  "Foto 3x4",
    "historico_escolar":         "Histórico Escolar",
    "certificado_conclusao":     "Certificado de Conclusão",
    "titulo_eleitor":            "Título de Eleitor",
    "certificado_reservista":    "Certificado de Reservista",
    "cartao_vacina_frente":      "Cartão de Vacina (Frente)",
    "cartao_vacina_verso":       "Cartão de Vacina (Verso)",
    "pis_doc":                   "Documento PIS/NIT",
}

REQUIRED_DOCUMENTS = [
    "rg_frente", "cpf_frente", "carteira_trabalho",
    "comprovante_residencia", "curriculo", "foto_3x4", "titulo_eleitor",
]

REQUIRED_FIELDS = {
    "nome_completo":           "Nome completo",
    "data_nascimento":         "Data de nascimento",
    "local_nascimento":        "Local de nascimento",
    "estado_natal":            "Estado de nascimento",
    "estado_civil":            "Estado civil",
    "nacionalidade":           "Nacionalidade",
    "escolaridade":            "Grau de escolaridade",
    "funcao":                  "Função/cargo",
    "nome_pai":                "Nome do pai",
    "nome_mae":                "Nome da mãe",
    "telefone":                "Telefone",
    "endereco":                "Endereço",
    "cep":                     "CEP",
    "bairro":                  "Bairro",
    "estado_residencia":       "Estado de residência",
    "numero_rg":               "Número do RG",
    "estado_emissor_rg":       "Estado emissor do RG",
    "data_emissao_rg":         "Data de emissão do RG",
    "numero_cpf":              "Número do CPF",
    "numero_carteira":         "Número da Carteira de Trabalho",
    "serie_carteira":          "Série da Carteira de Trabalho",
    "estado_emissor_carteira": "Estado emissor da Carteira",
    "data_emissao_carteira":   "Data de emissão da Carteira",
    "pis":                     "Número do PIS",
    "numero_titulo":           "Número do Título de Eleitor",
    "zona_eleitoral":          "Zona eleitoral",
    "secao_eleitoral":         "Seção eleitoral",
}

OPTIONAL_FIELDS = {
    "numero_reservista":       "Número do Certificado de Reservista",
    "serie_reservista":        "Série do Reservista",
    "categoria_reservista":    "Categoria do Reservista",
    "empresa_anterior":        "Empresa anterior",
    "cargo_anterior":          "Cargo anterior",
    "data_admissao_anterior":  "Data de admissão anterior",
    "data_demissao_anterior":  "Data de demissão anterior",
    "motivo_saida":            "Motivo da saída",
    "banco":                   "Banco",
    "agencia":                 "Agência",
    "conta":                   "Conta corrente",
}

CELL_MAP = {
    "nome_completo":           ("A3",  lambda v: f"NOME: {str(v).upper()}"),
    "endereco":                ("B4",  lambda v: str(v).upper()),
    "cep":                     ("A5",  lambda v: f"CEP:         {v}"),
    "bairro":                  ("C5",  lambda v: str(v).upper()),
    "estado_residencia":       ("F5",  lambda v: str(v).upper()),
    "telefone":                ("B6",  None),
    "data_nascimento":         ("A8",  None),
    "local_nascimento":        ("C8",  lambda v: str(v).upper()),
    "estado_natal":            ("F8",  lambda v: str(v).upper()),
    "estado_civil":            ("A10", lambda v: str(v).upper()),
    "nacionalidade":           ("B10", lambda v: str(v).upper()),
    "escolaridade":            ("D10", lambda v: str(v).upper()),
    "funcao":                  ("B11", lambda v: str(v).upper()),
    "nome_pai":                ("A14", lambda v: f"PAI: {str(v).upper()}"),
    "nome_mae":                ("A15", lambda v: f"MÃE:  {str(v).upper()}"),
    "numero_carteira":         ("H7",  None),
    "serie_carteira":          ("J7",  None),
    "estado_emissor_carteira": ("H9",  lambda v: str(v).upper()),
    "data_emissao_carteira":   ("J9",  None),
    "numero_rg":               ("H12", None),
    "estado_emissor_rg":       ("H14", lambda v: str(v).upper()),
    "data_emissao_rg":         ("J14", None),
    "numero_cpf":              ("H17", None),
    "numero_titulo":           ("H20", None),
    "zona_eleitoral":          ("H22", None),
    "secao_eleitoral":         ("J22", None),
    "numero_reservista":       ("H25", None),
    "serie_reservista":        ("H27", None),
    "categoria_reservista":    ("J27", None),
    "pis":                     ("H30", None),
    "empresa_anterior":        ("B23", lambda v: str(v).upper()),
    "endereco_empresa":        ("B24", lambda v: str(v).upper()),
    "cargo_anterior":          ("B25", lambda v: str(v).upper()),
    "data_admissao_anterior":  ("B26", None),
    "data_demissao_anterior":  ("D26", None),
    "motivo_saida":            ("B27", lambda v: str(v).upper()),
    "banco":                   ("B32", lambda v: str(v).upper()),
    "agencia":                 ("E32", None),
    "conta":                   ("B33", None),
    "tempo_experiencia":       ("B13", lambda v: f"Tempo: {v}"),
}

FIELD_PRIORITY = {
    "nome_completo":           ["rg_frente", "carteira_trabalho", "cpf_frente", "cnh", "curriculo"],
    "data_nascimento":         ["rg_frente", "carteira_trabalho", "cnh", "cpf_frente"],
    "local_nascimento":        ["rg_frente", "carteira_trabalho", "cnh"],
    "estado_natal":            ["rg_frente", "carteira_trabalho", "cnh"],
    "nome_pai":                ["rg_frente", "carteira_trabalho", "cnh"],
    "nome_mae":                ["rg_frente", "carteira_trabalho", "cnh"],
    "numero_rg":               ["rg_frente"],
    "estado_emissor_rg":       ["rg_frente"],
    "data_emissao_rg":         ["rg_frente"],
    "numero_cpf":              ["cpf_frente", "rg_verso", "carteira_trabalho_digital", "cnh"],
    "numero_carteira":         ["carteira_trabalho"],
    "serie_carteira":          ["carteira_trabalho"],
    "estado_emissor_carteira": ["carteira_trabalho"],
    "data_emissao_carteira":   ["carteira_trabalho"],
    "pis":                     ["pis_doc", "carteira_trabalho", "carteira_trabalho_digital"],
    "numero_titulo":           ["titulo_eleitor"],
    "zona_eleitoral":          ["titulo_eleitor"],
    "secao_eleitoral":         ["titulo_eleitor"],
    "numero_reservista":       ["certificado_reservista"],
    "serie_reservista":        ["certificado_reservista"],
    "categoria_reservista":    ["certificado_reservista"],
    "endereco":                ["comprovante_residencia"],
    "cep":                     ["comprovante_residencia"],
    "bairro":                  ["comprovante_residencia"],
    "estado_residencia":       ["comprovante_residencia"],
    "telefone":                ["curriculo"],
    "funcao":                  ["curriculo"],
    "escolaridade":            ["historico_escolar", "certificado_conclusao", "curriculo"],
    "empresa_anterior":        ["carteira_trabalho_digital", "curriculo"],
    "cargo_anterior":          ["carteira_trabalho_digital", "curriculo"],
    "data_admissao_anterior":  ["carteira_trabalho_digital", "curriculo"],
    "data_demissao_anterior":  ["carteira_trabalho_digital", "curriculo"],
    "motivo_saida":            ["carteira_trabalho_digital", "curriculo"],
    "endereco_empresa":        ["curriculo"],
    "estado_civil":            ["carteira_trabalho"],
    "nacionalidade":           ["carteira_trabalho", "rg_frente"],
    "tempo_experiencia":       ["curriculo"],
}

# ============================================================
# PROMPTS DE EXTRAÇÃO
# ============================================================

EXTRACTION_PROMPTS = {
    "rg_frente": """Analise esta imagem do RG (Registro Geral) brasileiro - frente.
Retorne APENAS JSON válido com estes campos (null se não visível):
{
  "nome_completo": "nome completo",
  "data_nascimento": "DD/MM/AAAA",
  "local_nascimento": "cidade",
  "estado_natal": "sigla UF ex: SP",
  "nome_pai": "nome do pai",
  "nome_mae": "nome da mãe",
  "numero_rg": "número do RG",
  "orgao_emissor": "ex: SSP",
  "estado_emissor_rg": "sigla UF",
  "data_emissao_rg": "DD/MM/AAAA"
}""",

    "rg_verso": """Analise o verso deste RG brasileiro.
Retorne APENAS JSON válido:
{
  "cpf_no_rg": "CPF se impresso",
  "profissao": "profissão se indicada"
}""",

    "cpf_frente": """Analise este documento CPF brasileiro.
Retorne APENAS JSON válido:
{
  "numero_cpf": "CPF no formato XXX.XXX.XXX-XX",
  "nome_completo": "nome completo",
  "data_nascimento": "DD/MM/AAAA"
}""",

    "cpf_verso": """Analise o verso deste CPF.
Retorne APENAS JSON válido:
{"informacoes_adicionais": "qualquer informação relevante"}""",

    "carteira_trabalho": """Analise esta Carteira de Trabalho e Previdência Social (CTPS) brasileira.
Pode ser a página de identificação ou de contratos. Retorne APENAS JSON válido:
{
  "numero_carteira": "número da CTPS",
  "serie_carteira": "série",
  "estado_emissor_carteira": "sigla UF",
  "data_emissao_carteira": "DD/MM/AAAA",
  "nome_completo": "nome completo",
  "data_nascimento": "DD/MM/AAAA",
  "local_nascimento": "cidade",
  "estado_natal": "sigla UF",
  "nome_pai": "nome do pai",
  "nome_mae": "nome da mãe",
  "nacionalidade": "nacionalidade",
  "estado_civil": "estado civil",
  "pis": "PIS/NIT se presente",
  "empresa_anterior": "última empresa registrada",
  "cargo_anterior": "último cargo",
  "data_admissao_anterior": "DD/MM/AAAA",
  "data_demissao_anterior": "DD/MM/AAAA ou null"
}""",

    "carteira_trabalho_digital": """Analise esta Carteira de Trabalho Digital brasileira.
Retorne APENAS JSON válido:
{
  "nome_completo": "nome completo",
  "data_nascimento": "DD/MM/AAAA",
  "numero_cpf": "CPF",
  "pis": "PIS/NIT",
  "estado_civil": "estado civil",
  "empresa_anterior": "empresa mais recente",
  "cargo_anterior": "cargo mais recente",
  "data_admissao_anterior": "DD/MM/AAAA",
  "data_demissao_anterior": "DD/MM/AAAA ou null",
  "motivo_saida": "motivo da saída ou null"
}""",

    "cnh": """Analise esta CNH (Carteira Nacional de Habilitação) brasileira.
Retorne APENAS JSON válido:
{
  "nome_completo": "nome completo",
  "numero_cpf": "CPF",
  "data_nascimento": "DD/MM/AAAA",
  "local_nascimento": "cidade",
  "estado_natal": "sigla UF",
  "nome_pai": "nome do pai",
  "nome_mae": "nome da mãe",
  "validade": "DD/MM/AAAA",
  "categoria": "categoria"
}""",

    "comprovante_residencia": """Analise este comprovante de residência (conta de luz/água/gás/telefone/banco).
Retorne APENAS JSON válido:
{
  "nome_titular": "nome do titular",
  "endereco_logradouro": "rua/avenida e número apenas",
  "complemento": "complemento ou null",
  "bairro": "bairro",
  "cidade": "cidade",
  "estado_residencia": "sigla UF",
  "cep": "CEP"
}""",

    "curriculo": """Analise este currículo e extraia as informações principais.
Retorne APENAS JSON válido:
{
  "nome_completo": "nome completo",
  "telefone": "telefone com DDD",
  "email": "e-mail",
  "funcao": "cargo/função em maiúsculas",
  "escolaridade": "ENSINO FUNDAMENTAL, ENSINO MÉDIO ou ENSINO SUPERIOR",
  "empresa_anterior": "empresa mais recente",
  "cargo_anterior": "cargo mais recente",
  "data_admissao_anterior": "DD/MM/AAAA",
  "data_demissao_anterior": "DD/MM/AAAA ou null",
  "endereco_empresa": "endereço da empresa anterior",
  "motivo_saida": "motivo da saída ou null",
  "tempo_experiencia": "ex: 5 anos, 8 meses"
}""",

    "foto_3x4": """Esta é uma foto 3x4 do candidato.
Retorne APENAS: {"foto_recebida": true}""",

    "historico_escolar": """Analise este histórico escolar.
Retorne APENAS JSON válido:
{
  "nome_aluno": "nome",
  "nivel_ensino": "ENSINO FUNDAMENTAL, ENSINO MÉDIO ou ENSINO SUPERIOR",
  "situacao": "CONCLUÍDO, CURSANDO ou INCOMPLETO",
  "ano_conclusao": "ano ou null",
  "instituicao": "nome da escola"
}""",

    "certificado_conclusao": """Analise este certificado de conclusão.
Retorne APENAS JSON válido:
{
  "nome_formando": "nome completo",
  "nivel_ensino": "ENSINO FUNDAMENTAL, ENSINO MÉDIO ou ENSINO SUPERIOR",
  "data_conclusao": "ano ou DD/MM/AAAA",
  "instituicao": "nome da instituição"
}""",

    "titulo_eleitor": """Analise este Título de Eleitor brasileiro.
Retorne APENAS JSON válido:
{
  "nome_completo": "nome completo",
  "data_nascimento": "DD/MM/AAAA",
  "numero_titulo": "número do título somente dígitos",
  "zona_eleitoral": "número da zona somente dígitos",
  "secao_eleitoral": "número da seção somente dígitos",
  "municipio": "município",
  "estado_titulo": "sigla UF"
}""",

    "certificado_reservista": """Analise este Certificado de Reservista (serviço militar brasileiro).
Retorne APENAS JSON válido:
{
  "nome_completo": "nome completo",
  "data_nascimento": "DD/MM/AAAA",
  "numero_reservista": "número do certificado",
  "serie_reservista": "série",
  "categoria_reservista": "categoria ex: 1ª CATEGORIA"
}""",

    "cartao_vacina_frente": """Analise a frente deste cartão de vacinação.
Retorne APENAS JSON válido:
{
  "nome_paciente": "nome completo",
  "data_nascimento": "DD/MM/AAAA",
  "cartao_recebido": true
}""",

    "cartao_vacina_verso": """Analise o verso deste cartão de vacinação.
Retorne APENAS JSON válido:
{"cartao_verso_recebido": true}""",

    "pis_doc": """Analise este documento brasileiro que contém o número do PIS/NIT/PASEP.
Pode ser cartão PIS, extrato CAIXA, comprovante PIS ou qualquer documento com o número PIS.
Retorne APENAS JSON válido:
{
  "pis": "número PIS/NIT completo com pontos e traço ex: 123.45678.12-3",
  "nome_completo": "nome do titular se visível"
}""",
}

# ============================================================
# BUSCA DE DOCUMENTOS
# ============================================================

def find_documents(folder: str) -> dict:
    found = {}
    folder_path = Path(folder)
    extensions = [".jpg", ".jpeg", ".png", ".pdf", ".webp"]
    for doc_type in DOCUMENT_TYPES:
        for ext in extensions:
            fp = folder_path / f"{doc_type}{ext}"
            if fp.exists():
                found[doc_type] = str(fp)
                break
    return found

# ============================================================
# EXTRAÇÃO VIA GEMINI
# ============================================================

MIME_TYPES = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".png": "image/png",  ".webp": "image/webp",
    ".gif": "image/gif",  ".pdf":  "application/pdf",
}


def parse_json_response(text: str) -> dict:
    text = text.strip()
    if "```" in text:
        start = text.find("```")
        end = text.rfind("```")
        inner = text[start + 3:end]
        if inner.startswith("json"):
            inner = inner[4:]
        text = inner.strip()
    return json.loads(text)


def extract_document_info(key_manager: GeminiKeyManager, doc_type: str, file_path: str) -> dict:
    prompt = EXTRACTION_PROMPTS.get(doc_type)
    if not prompt:
        return {}

    for attempt in range(3):
        api_key = key_manager.get_active_key()

        # Chave esgotada — pede nova interativamente
        if api_key is None:
            if not key_manager.request_new_key_interactive():
                print("    Processamento interrompido: sem chaves disponíveis.")
                return {}
            api_key = key_manager.get_active_key()
            if api_key is None:
                return {}

        try:
            client = genai.Client(api_key=api_key)

            suffix = Path(file_path).suffix.lower()
            mime_type = MIME_TYPES.get(suffix, "image/jpeg")

            with open(file_path, "rb") as f:
                file_bytes = f.read()

            part = gtypes.Part.from_bytes(data=file_bytes, mime_type=mime_type)
            response = client.models.generate_content(
                model="gemini-flash-lite-latest",
                contents=[part, prompt],
            )

            key_manager.consume()
            key_manager.warn_if_low()

            extracted = parse_json_response(response.text)
            return {k: v for k, v in extracted.items()
                    if v is not None and v != "" and v is not False}

        except Exception as e:
            err = str(e).lower()
            is_quota = "quota" in err or "429" in err or "resource_exhausted" in err
            is_auth  = "api_key" in err or "invalid" in err or "401" in err or "403" in err or "api key" in err
            if is_quota or is_auth:
                reason = "Cota excedida" if is_quota else "Chave inválida/sem permissão"
                print(f"    [!] {reason}. Trocando de chave...")
                entry = key_manager._current_entry()
                if entry:
                    entry["requests_today"] = DAILY_LIMIT
                    key_manager._save()
                continue
            if attempt == 2:
                print(f"    Erro ao processar {doc_type}: {e}")
                raise
            else:
                print(f"    Tentativa {attempt + 1} falhou: {e}")

    return {}

# ============================================================
# CONSOLIDAÇÃO DE DADOS
# ============================================================

def merge_extracted_data(all_extractions: dict) -> dict:
    merged = {}
    aliases = {
        "cpf_no_rg":           "numero_cpf",
        "nome_titular":        "nome_completo",
        "nome_paciente":       "nome_completo",
        "nome_aluno":          "nome_completo",
        "nome_formando":       "nome_completo",
        "endereco_logradouro": "endereco",
        "nivel_ensino":        "escolaridade",
    }

    all_fields = (
        list(REQUIRED_FIELDS) + list(OPTIONAL_FIELDS) +
        ["tempo_experiencia", "endereco_empresa", "email"]
    )

    for field in all_fields:
        priority_docs = FIELD_PRIORITY.get(field, list(DOCUMENT_TYPES))
        for doc_type in priority_docs:
            if doc_type not in all_extractions:
                continue
            extraction = all_extractions[doc_type]
            if field in extraction and extraction[field]:
                merged[field] = extraction[field]
                break
            for alias, canonical in aliases.items():
                if canonical == field and alias in extraction and extraction[alias]:
                    merged[field] = extraction[alias]
                    break
            if field in merged:
                break

    if "curriculo" in all_extractions:
        c = all_extractions["curriculo"]
        for f in ("email", "tempo_experiencia", "endereco_empresa"):
            if f in c and c[f] and f not in merged:
                merged[f] = c[f]

    return merged

# ============================================================
# PREENCHIMENTO DO EXCEL
# ============================================================

def calculate_age(birth_str: str) -> Optional[int]:
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            bd = datetime.strptime(str(birth_str), fmt)
            today = datetime.today()
            return today.year - bd.year - (
                (today.month, today.day) < (bd.month, bd.day)
            )
        except ValueError:
            continue
    return None


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return value


def fill_excel(template_path: str, output_path: str, data: dict) -> list:
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for field, (cell_coord, fmt_func) in CELL_MAP.items():
        value = data.get(field)
        if not value:
            continue
        if fmt_func:
            cell_value = fmt_func(value)
        elif isinstance(value, str) and "/" in value and 5 <= len(value) <= 10:
            # Normaliza para DD/MM/AAAA como string (evita exibição datetime no Excel)
            parsed = parse_date(value)
            if isinstance(parsed, datetime):
                cell_value = parsed.strftime("%d/%m/%Y")
            else:
                cell_value = value
        else:
            cell_value = value
        ws[cell_coord] = cell_value

    if "data_nascimento" in data:
        age = calculate_age(str(data["data_nascimento"]))
        if age:
            ws["B8"] = age

    ws["H4"] = MESES_PT[datetime.today().month]
    wb.save(output_path)

    missing = [
        (field, label)
        for field, label in REQUIRED_FIELDS.items()
        if not data.get(field)
    ]
    return missing

# ============================================================
# RELATÓRIO
# ============================================================

def generate_report(found_docs, merged_data, missing_fields, output_excel, key_manager) -> str:
    lines = []
    sep = "=" * 65
    lines += [
        sep,
        "   RELATÓRIO DE SOLICITAÇÃO DE EMPREGO - FIBRA",
        sep,
        f"Data/Hora : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Candidato : {str(merged_data.get('nome_completo', 'Não identificado')).upper()}",
        "",
    ]

    lines.append("DOCUMENTOS RECEBIDOS:")
    lines.append("-" * 50)
    for doc_type, display_name in DOCUMENT_TYPES.items():
        ok = doc_type in found_docs
        req = "(obrigatório)" if doc_type in REQUIRED_DOCUMENTS else "(opcional)  "
        icon = "[OK]" if ok else "[--]"
        lines.append(f"  {icon}  {display_name:<40} {req}")
    lines.append("")

    faltando_docs = [DOCUMENT_TYPES[d] for d in REQUIRED_DOCUMENTS if d not in found_docs]
    if faltando_docs:
        lines.append("ATENÇÃO - DOCUMENTOS OBRIGATÓRIOS FALTANDO:")
        lines.append("-" * 50)
        for doc in faltando_docs:
            lines.append(f"  *** {doc}")
        lines.append("")

    if missing_fields:
        lines.append("CAMPOS NÃO PREENCHIDOS (INFORMAÇÕES AUSENTES):")
        lines.append("-" * 50)
        for _, label in missing_fields:
            lines.append(f"  *** {label}")
        lines.append("")
        lines.append("AÇÃO: Solicite os documentos/informações faltando e")
        lines.append("      preencha manualmente no Excel.")
    else:
        lines.append("RESULTADO: Todos os campos obrigatórios foram preenchidos!")
    lines.append("")

    lines.append("DADOS EXTRAÍDOS:")
    lines.append("-" * 50)
    for field, label in {**REQUIRED_FIELDS, **OPTIONAL_FIELDS}.items():
        val = merged_data.get(field)
        if val:
            lines.append(f"  [OK] {label}: {val}")
    lines.append("")

    lines.append("USO DAS CHAVES GEMINI:")
    lines.append("-" * 50)
    for line in key_manager.print_status.__wrapped__(key_manager) if hasattr(key_manager.print_status, '__wrapped__') else []:
        lines.append(line)

    lines += [f"\nArquivo Excel: {output_excel}", sep]
    return "\n".join(lines)

# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Solicitação de Emprego FIBRA — preenche Excel a partir de documentos via Gemini",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("--documentos", "-d", required=True,
                        help="Pasta com os documentos do candidato")
    parser.add_argument("--template", "-t",
                        default=str(Path(__file__).parent / "template_solicitacao.xlsx"),
                        help="Caminho para o Excel template")
    parser.add_argument("--saida", "-s",
                        help="Arquivo Excel de saída (padrão: SOLICITACAO_[NOME].xlsx)")
    parser.add_argument("--api-key", dest="api_keys", action="append", default=[],
                        metavar="CHAVE",
                        help="Chave Gemini API (pode usar várias vezes para múltiplas contas)")
    parser.add_argument("--status", action="store_true",
                        help="Mostra o status das chaves cadastradas e sai")
    args = parser.parse_args()

    # Inicializa gerenciador de chaves
    key_manager = GeminiKeyManager()

    # Adiciona chaves fornecidas via argumento
    for key in args.api_keys:
        added = key_manager.add_key(key)
        if added:
            print(f"[OK] Chave ...{key[-8:]} adicionada.")

    # Tenta variável de ambiente como fallback
    env_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if env_key and not key_manager.has_keys():
        key_manager.add_key(env_key)

    if args.status:
        print("\nStatus das chaves Gemini:")
        key_manager.print_status()
        print(f"\nLimite diário: {DAILY_LIMIT} req/dia")
        print(f"Buffer reservado: {STOP_AT_REMAINING} req (não utilizadas)")
        return

    # Validações
    if not Path(args.documentos).is_dir():
        print(f"Erro: pasta não encontrada: {args.documentos}")
        sys.exit(1)

    if not Path(args.template).exists():
        print(f"Erro: template não encontrado: {args.template}")
        sys.exit(1)

    if not key_manager.has_keys():
        print("Erro: nenhuma chave Gemini configurada.")
        print("Use --api-key SUA_CHAVE ou defina a variável GEMINI_API_KEY.")
        print("Obtenha uma chave gratuita em: https://aistudio.google.com/app/apikey")
        sys.exit(1)

    print("\n" + "=" * 65)
    print("   SISTEMA DE SOLICITAÇÃO DE EMPREGO - FIBRA")
    print("=" * 65)
    print("\nChaves configuradas:")
    key_manager.print_status()

    # 1. Busca documentos
    print(f"\n[1/4] Buscando documentos em: {args.documentos}")
    found_docs = find_documents(args.documentos)
    print(f"      {len(found_docs)} documento(s) encontrado(s).")
    for doc_type, path in found_docs.items():
        print(f"        - {DOCUMENT_TYPES[doc_type]}: {Path(path).name}")

    if not found_docs:
        print("\nNenhum documento encontrado. Verifique a pasta e os nomes dos arquivos.")
        sys.exit(1)

    # 2. Extração
    print("\n[2/4] Extraindo informações via Gemini AI...")
    all_extractions = {}
    for doc_type, file_path in found_docs.items():
        display = DOCUMENT_TYPES[doc_type]
        print(f"      Processando: {display}...")
        extraction = extract_document_info(key_manager, doc_type, file_path)
        if extraction:
            all_extractions[doc_type] = extraction
            print(f"        -> {len(extraction)} campo(s) extraído(s)")
        else:
            print(f"        -> Nenhum dado extraído")

    # 3. Consolidação
    print("\n[3/4] Consolidando informações...")
    merged_data = merge_extracted_data(all_extractions)
    print(f"      {len(merged_data)} campo(s) prontos para o Excel.")

    # Define saída
    nome = str(merged_data.get("nome_completo", "candidato")).upper()
    nome_safe = nome.replace(" ", "_").replace("/", "-")[:50]
    output_path = args.saida or f"SOLICITACAO_{nome_safe}.xlsx"

    # 4. Preenchimento
    print(f"\n[4/4] Preenchendo Excel: {output_path}")
    missing_fields = fill_excel(args.template, output_path, merged_data)

    # Relatório
    report_lines = [
        "=" * 65,
        "   RELATÓRIO DE SOLICITAÇÃO DE EMPREGO - FIBRA",
        "=" * 65,
        f"Data/Hora : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Candidato : {nome}",
        "",
        "DOCUMENTOS RECEBIDOS:",
        "-" * 50,
    ]
    for doc_type, display_name in DOCUMENT_TYPES.items():
        ok = doc_type in found_docs
        req = "(obrigatório)" if doc_type in REQUIRED_DOCUMENTS else "(opcional)  "
        icon = "[OK]" if ok else "[--]"
        report_lines.append(f"  {icon}  {display_name:<40} {req}")
    report_lines.append("")

    faltando_docs = [DOCUMENT_TYPES[d] for d in REQUIRED_DOCUMENTS if d not in found_docs]
    if faltando_docs:
        report_lines += ["ATENÇÃO - DOCUMENTOS OBRIGATÓRIOS FALTANDO:", "-" * 50]
        report_lines += [f"  *** {d}" for d in faltando_docs]
        report_lines.append("")

    if missing_fields:
        report_lines += ["CAMPOS NÃO PREENCHIDOS:", "-" * 50]
        report_lines += [f"  *** {label}" for _, label in missing_fields]
        report_lines.append("")
    else:
        report_lines.append("RESULTADO: Todos os campos obrigatórios foram preenchidos!")
        report_lines.append("")

    report_lines += ["DADOS EXTRAÍDOS:", "-" * 50]
    for field, label in {**REQUIRED_FIELDS, **OPTIONAL_FIELDS}.items():
        val = merged_data.get(field)
        if val:
            report_lines.append(f"  [OK] {label}: {val}")
    report_lines += ["", "USO DAS CHAVES GEMINI:", "-" * 50]

    report_text = "\n".join(report_lines)
    report_path = output_path.replace(".xlsx", "_relatorio.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_text)

    print("\n" + report_text)
    print(f"\nRelatório salvo em: {report_path}")
    print("\nStatus final das chaves:")
    key_manager.print_status()

    if missing_fields:
        print(f"\n[!] {len(missing_fields)} campo(s) precisam ser preenchidos manualmente.")
        sys.exit(2)
    else:
        print("\n[OK] Processamento concluído com sucesso!")
        sys.exit(0)


if __name__ == "__main__":
    main()
